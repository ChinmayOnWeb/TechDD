from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from acquirescope.bridge.adjustments import Adjustment
from acquirescope.bridge.assumptions import Assumptions
from acquirescope.bridge.valuation import DcfScenario, SensitivityGrid, Valuation
from acquirescope.report import DISCLAIMER

_FONT_NAME = "Calibri"
_TITLE_FONT = Font(name=_FONT_NAME, size=14, bold=True, color="1F3864")
_SECTION_FONT = Font(name=_FONT_NAME, size=11, bold=True, color="1F3864")
_HEADER_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_TOTAL_FONT = Font(name=_FONT_NAME, size=10, bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
# Industry-standard model color coding: blue = hardcoded input, black = formula
# or Python-computed value, green = a formula linking to another sheet.
_INPUT_FONT = Font(name=_FONT_NAME, size=10, color="0000FF")
_CALC_FONT = Font(name=_FONT_NAME, size=10, color="000000")
_LINK_FONT = Font(name=_FONT_NAME, size=10, color="008000")
_MUTED_FONT = Font(name=_FONT_NAME, size=10, italic=True, color="808080")
_OK_FONT = Font(name=_FONT_NAME, size=10, color="008000")
_NOT_ASSESSED_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="C00000")
_BOLD = Font(name=_FONT_NAME, size=10, bold=True, color="1F3864")

_THIN = Side(style="thin", color="B7B7B7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

_USD_FMT = '$#,##0;($#,##0);"-"'
_PCT_FMT = '0.0%;(0.0%);"-"'
_MULT_FMT = '0.0"x"'


def _rows(ws: Worksheet, start_row: int, rows: list[list]) -> int:
    """Write rows starting at start_row; returns the next free row."""
    for r, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    return start_row + len(rows)


def _set_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _box(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _title(ws: Worksheet, cell: str, text: str, span: int = 4) -> None:
    ws[cell] = text
    ws[cell].font = _TITLE_FONT
    row = ws[cell].row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _section(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = _SECTION_FONT


def _sheet_assumptions(ws: Worksheet, a: Assumptions) -> None:
    _title(ws, "A1", f"Assumptions: {a.target_name}", span=3)
    header_row = 3
    data = [
        ("Revenue low (USD/yr)", a.revenue_low, a.revenue_source, _USD_FMT),
        ("Revenue mid (USD/yr)", a.revenue_mid, a.revenue_source, _USD_FMT),
        ("Revenue high (USD/yr)", a.revenue_high, a.revenue_source, _USD_FMT),
        ("DCF years", a.dcf_years, "", "0"),
        ("Growth bear", a.growth_bear, "", _PCT_FMT),
        ("Growth base", a.growth_base, "", _PCT_FMT),
        ("Growth bull", a.growth_bull, "", _PCT_FMT),
        ("Operating margin", a.operating_margin, "", _PCT_FMT),
        ("Discount rate", a.discount_rate, "", _PCT_FMT),
        ("Terminal growth", a.terminal_growth, "", _PCT_FMT),
        ("Engineer-month cost (USD)", a.engineer_month_usd, "", _USD_FMT),
        ("Retention package (USD)", a.retention_package_usd, "", _USD_FMT),
        ("Integration cost (USD)", a.integration_cost_usd, "", _USD_FMT),
        ("Security fix cost (USD)", a.security_fix_cost_usd, "", _USD_FMT),
        ("License discount per finding", a.license_discount_per_finding, "", _PCT_FMT),
        ("License discount cap", a.license_discount_cap, "", _PCT_FMT),
    ]
    _rows(ws, header_row, [["Input", "Value", "Source"]])
    _style_header_row(ws, header_row, 3)
    last_row = _rows(ws, header_row + 1, [[label, value, source] for label, value, source, _ in data]) - 1
    for i, (_, _, _, fmt) in enumerate(data):
        r = header_row + 1 + i
        ws.cell(row=r, column=2).font = _INPUT_FONT
        ws.cell(row=r, column=2).number_format = fmt
        ws.cell(row=r, column=3).alignment = _WRAP
    _box(ws, header_row, last_row, 1, 3)
    _set_widths(ws, {"A": 30, "B": 16, "C": 75})
    ws.freeze_panes = "A4"


def _sheet_comps(ws: Worksheet, a: Assumptions, comps: Valuation) -> None:
    _title(ws, "A1", "Comparables")
    ws["A2"] = f"Source: {a.comps_source}"
    ws["A2"].alignment = _WRAP
    ws.merge_cells("A2:B2")
    header_row = 4
    _rows(ws, header_row, [["Company", "EV/Revenue multiple"]])
    _style_header_row(ws, header_row, 2)
    last_row = _rows(ws, header_row + 1, [[c.name, c.ev_revenue_multiple] for c in a.comps]) - 1
    for r in range(header_row + 1, last_row + 1):
        ws.cell(row=r, column=2).font = _INPUT_FONT
        ws.cell(row=r, column=2).number_format = _MULT_FMT
    _box(ws, header_row, last_row, 1, 2)

    ev_header_row = last_row + 2
    _rows(ws, ev_header_row, [["Implied EV (median multiple)", "Low", "Mid", "High"]])
    _style_header_row(ws, ev_header_row, 4)
    ev_row = ev_header_row + 1
    _rows(ws, ev_row, [["", comps.low, comps.mid, comps.high]])
    for c in range(2, 5):
        cell = ws.cell(row=ev_row, column=c)
        cell.font = _CALC_FONT
        cell.number_format = _USD_FMT
    _box(ws, ev_header_row, ev_row, 1, 4)
    _set_widths(ws, {"A": 30, "B": 18, "C": 18, "D": 18})
    ws.freeze_panes = "A5"


def _sheet_dcf(ws: Worksheet, scenarios: list[DcfScenario]) -> None:
    _title(ws, "A1", "DCF — bear / base / bull scenarios")
    row = 3
    for s in scenarios:
        _section(ws, f"A{row}", f"Scenario: {s.name} (growth {s.growth:.0%})")
        header_row = row + 1
        _rows(ws, header_row, [["Year", "Revenue", "FCF", "PV"]])
        _style_header_row(ws, header_row, 4)
        data_start = header_row + 1
        data_end = _rows(ws, data_start, [[y.year, y.revenue, y.fcf, y.pv] for y in s.years]) - 1
        for r in range(data_start, data_end + 1):
            ws.cell(row=r, column=1).number_format = "0"
            for c in (2, 3, 4):
                cell = ws.cell(row=r, column=c)
                cell.font = _CALC_FONT
                cell.number_format = _USD_FMT
        _box(ws, header_row, data_end, 1, 4)

        summary_row = data_end + 1
        row = _rows(ws, summary_row, [["Terminal PV", s.terminal_pv], ["Scenario EV", s.ev]])
        for r in (summary_row, summary_row + 1):
            ws.cell(row=r, column=1).font = _BOLD
            ws.cell(row=r, column=2).font = _CALC_FONT
            ws.cell(row=r, column=2).number_format = _USD_FMT
        row += 1
    _set_widths(ws, {"A": 34, "B": 18, "C": 18, "D": 18})


def _sheet_adjustments(ws: Worksheet, adjustments: list[Adjustment]) -> int:
    """Returns the total row index (Valuation Summary references it)."""
    header_row = 1
    _rows(ws, header_row, [["Line item", "Low (USD)", "Mid (USD)", "High (USD)", "Assessed", "Basis", "Evidence"]])
    _style_header_row(ws, header_row, 7)
    data_start = header_row + 1
    data_end = _rows(ws, data_start, [
        [adj.name, adj.low, adj.mid, adj.high,
         "yes" if adj.assessed else "NOT ASSESSED", adj.basis, "; ".join(adj.evidence)]
        for adj in adjustments
    ]) - 1
    for r, adj in zip(range(data_start, data_end + 1), adjustments):
        for c in (2, 3, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = _CALC_FONT
            cell.number_format = _USD_FMT
        assessed_cell = ws.cell(row=r, column=5)
        assessed_cell.font = _OK_FONT if adj.assessed else _NOT_ASSESSED_FONT
        assessed_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).alignment = _WRAP
        ws.cell(row=r, column=7).alignment = _WRAP
        if not adj.assessed:
            for c in range(1, 8):
                if c != 5:
                    ws.cell(row=r, column=c).font = _MUTED_FONT

    total_row = data_end + 1
    last = total_row - 1
    _rows(ws, total_row, [[
        "Total", f"=SUM(B2:B{last})", f"=SUM(C2:C{last})", f"=SUM(D2:D{last})",
    ]])
    for c in range(1, 5):
        cell = ws.cell(row=total_row, column=c)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        if c > 1:
            cell.number_format = _USD_FMT
    _box(ws, header_row, total_row, 1, 7)
    _set_widths(ws, {"A": 26, "B": 15, "C": 15, "D": 15, "E": 13, "F": 55, "G": 55})
    ws.freeze_panes = "A2"
    return total_row


def _sheet_summary(
    ws: Worksheet, a: Assumptions, comps: Valuation, dcf: Valuation,
    grid: SensitivityGrid, adjustments_total_row: int,
) -> None:
    _title(ws, "A1", f"Valuation Summary: {a.target_name}")
    t = adjustments_total_row
    header_row = 3
    _rows(ws, header_row, [["Method", "Low", "Mid", "High"]])
    _style_header_row(ws, header_row, 4)
    _rows(ws, header_row + 1, [
        ["Comps EV", comps.low, comps.mid, comps.high],
        ["DCF EV", dcf.low, dcf.mid, dcf.high],
        ["Blended pre-DD EV", "=AVERAGE(B4:B5)", "=AVERAGE(C4:C5)", "=AVERAGE(D4:D5)"],
    ])
    for r in (4, 5):
        for c in (2, 3, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = _CALC_FONT
            cell.number_format = _USD_FMT
    for c in (2, 3, 4):
        cell = ws.cell(row=6, column=c)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.number_format = _USD_FMT
    ws.cell(row=6, column=1).font = _TOTAL_FONT
    ws.cell(row=6, column=1).fill = _TOTAL_FILL
    _box(ws, header_row, 6, 1, 4)

    _rows(ws, 8, [
        ["Total DD adjustments",
         f"='DD Adjustments'!B{t}", f"='DD Adjustments'!C{t}", f"='DD Adjustments'!D{t}"],
        ["Post-DD EV", "=B6-B8", "=C6-C8", "=D6-D8"],
    ])
    for c in (2, 3, 4):
        ws.cell(row=8, column=c).font = _LINK_FONT
        ws.cell(row=8, column=c).number_format = _USD_FMT
        ws.cell(row=9, column=c).font = _TOTAL_FONT
        ws.cell(row=9, column=c).fill = _TOTAL_FILL
        ws.cell(row=9, column=c).number_format = _USD_FMT
    ws.cell(row=9, column=1).font = _TOTAL_FONT
    ws.cell(row=9, column=1).fill = _TOTAL_FILL
    _box(ws, 8, 9, 1, 4)

    _section(ws, "A11", "Sensitivity — Pre-DD EV (multiple x revenue)")
    _rows(ws, 12, [["Multiple \\ Revenue", *grid.revenues]])
    _style_header_row(ws, 12, 4)
    for c in (2, 3, 4):
        ws.cell(row=12, column=c).number_format = _USD_FMT
    for i, multiple in enumerate(grid.multiples):
        r = 13 + i
        ws.cell(row=r, column=1, value=multiple).number_format = _MULT_FMT
        ws.cell(row=r, column=1).font = _INPUT_FONT
        for c in range(2, 5):
            col = chr(ord("A") + c - 1)
            cell = ws.cell(row=r, column=c, value=f"=$A{r}*{col}$12")
            cell.font = _CALC_FONT
            cell.number_format = _USD_FMT
    _box(ws, 12, 15, 1, 4)

    _section(ws, "A17", "Sensitivity — Post-DD EV (pre-DD minus mid adjustments)")
    _rows(ws, 18, [["Multiple \\ Revenue", *grid.revenues]])
    _style_header_row(ws, 18, 4)
    for c in (2, 3, 4):
        ws.cell(row=18, column=c).number_format = _USD_FMT
    for i, multiple in enumerate(grid.multiples):
        r = 19 + i
        ws.cell(row=r, column=1, value=multiple).number_format = _MULT_FMT
        ws.cell(row=r, column=1).font = _INPUT_FONT
        for c in range(2, 5):
            col = chr(ord("A") + c - 1)
            cell = ws.cell(row=r, column=c, value=f"={col}{13 + i}-$C$8")
            cell.font = _CALC_FONT
            cell.number_format = _USD_FMT
    _box(ws, 18, 21, 1, 4)

    ws["A23"] = DISCLAIMER.strip("*")
    ws["A23"].font = _MUTED_FONT
    ws["A23"].alignment = _WRAP
    ws.merge_cells("A23:D25")
    _set_widths(ws, {"A": 24, "B": 18, "C": 18, "D": 18})
    ws.freeze_panes = "A3"


def write_model(
    path: Path,
    assumptions: Assumptions,
    adjustments: list[Adjustment],
    comps: Valuation,
    dcf: Valuation,
    scenarios: list[DcfScenario],
    grid: SensitivityGrid,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    _sheet_assumptions(ws, assumptions)
    _sheet_comps(wb.create_sheet("Comps"), assumptions, comps)
    _sheet_dcf(wb.create_sheet("DCF"), scenarios)
    total_row = _sheet_adjustments(wb.create_sheet("DD Adjustments"), adjustments)
    _sheet_summary(wb.create_sheet("Valuation Summary"), assumptions, comps, dcf, grid, total_row)
    wb.save(path)
