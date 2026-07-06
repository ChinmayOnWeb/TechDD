import json
from pathlib import Path

from typer.testing import CliRunner

from git_due_diligence import cli

runner = CliRunner()


def test_analyze_writes_report(fixture_repo, tmp_path):
    out = tmp_path / "report.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out)])
    assert result.exit_code == 0
    assert out.exists()
    md = out.read_text(encoding="utf-8")
    assert "# Technical Due Diligence Report:" in md


def test_module_crash_degrades_gracefully(fixture_repo, tmp_path, monkeypatch):
    def explode(ingest):
        raise RuntimeError("boom")

    monkeypatch.setattr(cli, "MODULES", [("hotspots", explode)] + [m for m in cli.MODULES if m[0] != "hotspots"])
    out = tmp_path / "report.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out)])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    assert "Not assessed" in md
    assert "boom" in md
    # other modules still ran
    assert "Module: bus_factor" in md


def test_all_planted_issues_detected_end_to_end(fixture_repo, tmp_path, monkeypatch):
    """Spec regression gate: every planted issue in the synthetic repo is found."""
    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    out = tmp_path / "e2e-report.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out)])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")

    # Planted issue 1: payments/ single-owner module
    assert "Single point of failure: payments/" in md
    # Planted issue 2: GPL dependency in requirements.txt
    assert "Copyleft dependency: mysqlclient" in md
    # Planted issue 3: churn+complexity hotspot
    assert "Tech-debt hotspot: core/engine.py" in md
    # Planted issue 4: inactive key contributor
    assert "Key contributor inactive: dave@example.com" in md
    # Confidence band present in metrics line
    assert "remediation_months_low" in md
    # Phase 3 planted issue: committed-then-removed AWS key
    assert "Secret in git history: AWS access key" in md
    # New modules render, honest about proxies and scope
    assert "## Module: security" in md
    assert "## Module: delivery" in md
    assert "Vulnerability scan not available" in md
    assert "No CI configuration detected" in md


def test_model_writes_workbook(fixture_repo, tmp_path):
    from openpyxl import load_workbook

    example = Path(__file__).parent.parent / "examples" / "assumptions.example.toml"
    out = tmp_path / "model.xlsx"
    result = runner.invoke(cli.app, [
        "model", str(fixture_repo), "--assumptions", str(example), "--output", str(out),
    ])
    assert result.exit_code == 0
    wb = load_workbook(out)
    assert "Valuation Summary" in wb.sheetnames


def test_model_invalid_assumptions_exits_1(fixture_repo, tmp_path):
    bad = tmp_path / "bad.toml"
    bad.write_text("[target]\nname = 'x'\n", encoding="utf-8")
    out = tmp_path / "model.xlsx"
    result = runner.invoke(cli.app, [
        "model", str(fixture_repo), "--assumptions", str(bad), "--output", str(out),
    ])
    assert result.exit_code == 1
    assert not out.exists()


def test_planted_issues_priced_in_model_end_to_end(fixture_repo, tmp_path, monkeypatch):
    """Phase 2 regression gate: planted DD issues become priced line items."""
    from openpyxl import load_workbook

    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    example = Path(__file__).parent.parent / "examples" / "assumptions.example.toml"
    out = tmp_path / "e2e-model.xlsx"
    result = runner.invoke(cli.app, [
        "model", str(fixture_repo), "--assumptions", str(example), "--output", str(out),
    ])
    assert result.exit_code == 0
    wb = load_workbook(out)

    adj = wb["DD Adjustments"]
    # Row order fixed: 2 remediation, 3 retention, 4 license, 5 integration, 6 security
    assert adj["B2"].value > 0                        # remediation capex priced from hotspot
    assert adj["C3"].value == 600_000                 # exactly 2 key-person findings x 300k
    assert adj["C4"].value > 0                        # license discount priced
    assert "mysqlclient" in adj["G4"].value           # evidence reaches the workbook
    assert adj["E6"].value == "yes"                   # security now assessed
    assert adj["C6"].value == 50_000                  # 1 CRITICAL secret x 50k
    assert "AWS access key" in adj["G6"].value        # evidence reaches the workbook

    summary = wb["Valuation Summary"]
    # Formulas are not evaluated by openpyxl; recompute from stored values.
    pre_mid = (summary["C4"].value + summary["C5"].value) / 2
    total_mid = sum(adj.cell(row=r, column=3).value for r in range(2, 7))
    assert total_mid > 0
    assert pre_mid - total_mid < pre_mid              # post-DD EV strictly below pre-DD


def test_narrative_unavailable_exits_1(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_llm_unavailable_reason", lambda: "no key")
    out = tmp_path / "r.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--narrative"])
    assert result.exit_code == 1
    assert not out.exists()


def test_narrative_section_written_with_fake_completer(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_llm_unavailable_reason", lambda: None)
    monkeypatch.setattr(cli, "_anthropic_complete",
                        lambda prompt: "Key-person risk dominates [E1]. Fake [E99].")
    out = tmp_path / "r.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--narrative"])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    assert "## Executive narrative" in md
    assert "[E1]" in md
    assert "[E99]" not in md
    assert "1 unverifiable citation(s) removed" in md


def test_narrative_api_failure_degrades_to_plain_report(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_llm_unavailable_reason", lambda: None)

    def boom(prompt):
        raise RuntimeError("api down")

    monkeypatch.setattr(cli, "_anthropic_complete", boom)
    out = tmp_path / "r.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--narrative"])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    assert "Executive narrative" not in md
    assert "# Technical Due Diligence Report:" in md


def test_dispositions_bootstrap_creates_pending_file(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    out = tmp_path / "report.md"
    disp_path = tmp_path / "dispositions.json"
    result = runner.invoke(cli.app, [
        "analyze", str(fixture_repo), "--output", str(out), "--dispositions", str(disp_path),
    ])
    assert result.exit_code == 0
    assert disp_path.exists()
    data = json.loads(disp_path.read_text(encoding="utf-8"))
    assert data["dispositions"]
    assert all(v["status"] == "pending" for v in data["dispositions"].values())


def test_malformed_dispositions_file_exits_1(fixture_repo, tmp_path):
    out = tmp_path / "report.md"
    disp_path = tmp_path / "bad.json"
    disp_path.write_text("{not valid", encoding="utf-8")
    result = runner.invoke(cli.app, [
        "analyze", str(fixture_repo), "--output", str(out), "--dispositions", str(disp_path),
    ])
    assert result.exit_code == 1
    assert not out.exists()


def test_dispositions_omitted_behaves_identically_to_before(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    out = tmp_path / "report.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out)])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    assert "Appendix: Dismissed" not in md


def _dismiss_dave_inactive_finding(disp_path: Path) -> None:
    data = json.loads(disp_path.read_text(encoding="utf-8"))
    target_id = next(
        fid for fid, v in data["dispositions"].items()
        if v["finding_title"] == "Key contributor inactive: dave@example.com"
    )
    data["dispositions"][target_id]["status"] = "dismissed"
    data["dispositions"][target_id]["note"] = "Confirmed still active via LinkedIn"
    disp_path.write_text(json.dumps(data), encoding="utf-8")


def test_dismissed_disposition_removes_finding_from_report(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    out = tmp_path / "report.md"
    disp_path = tmp_path / "dispositions.json"
    # First run: bootstrap the dispositions file (all pending).
    runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--dispositions", str(disp_path)])
    _dismiss_dave_inactive_finding(disp_path)

    # Second run: apply the analyst's dismissal.
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--dispositions", str(disp_path)])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    body, _, appendix = md.partition("## Appendix: Dismissed Findings")
    assert "Key contributor inactive: dave@example.com" not in body
    assert "Key contributor inactive: dave@example.com" in appendix
    assert "Confirmed still active via LinkedIn" in appendix


def test_dismissed_finding_excluded_from_model_pricing(fixture_repo, tmp_path, monkeypatch):
    from openpyxl import load_workbook

    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    example = Path(__file__).parent.parent / "examples" / "assumptions.example.toml"
    out = tmp_path / "report.md"
    disp_path = tmp_path / "dispositions.json"
    runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--dispositions", str(disp_path)])
    _dismiss_dave_inactive_finding(disp_path)

    xlsx_out = tmp_path / "model.xlsx"
    result = runner.invoke(cli.app, [
        "model", str(fixture_repo), "--assumptions", str(example),
        "--output", str(xlsx_out), "--dispositions", str(disp_path),
    ])
    assert result.exit_code == 0
    wb = load_workbook(xlsx_out)
    adj = wb["DD Adjustments"]
    # Fixture's bus_factor produces exactly 2 findings normally (payments/
    # single-owner + dave inactive); with dave dismissed, only 1 remains,
    # so retention prices at 1 x $300,000 instead of 2 x $300,000.
    assert adj["C3"].value == 300_000


def test_questions_unavailable_exits_1(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr(cli, "_llm_unavailable_reason", lambda: "no key")
    out = tmp_path / "r.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--questions"])
    assert result.exit_code == 1
    assert not out.exists()


def test_questions_rendered_with_fake_completer(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    monkeypatch.setattr(cli, "_llm_unavailable_reason", lambda: None)

    def fake_complete(prompt: str) -> str:
        import re
        real_id = re.search(r"\[([0-9a-f]{12})\] .*Single point of failure: payments/", prompt).group(1)
        return json.dumps({real_id: "Who owns payments/ if the current maintainer is unavailable?",
                            "bogus000000": "Should be dropped"})

    monkeypatch.setattr(cli, "_anthropic_complete", fake_complete)
    out = tmp_path / "r.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--questions"])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    assert "**Question for management:** Who owns payments/ if the current maintainer is unavailable?" in md
    assert "Should be dropped" not in md


def test_questions_failure_degrades_to_plain_report(fixture_repo, tmp_path, monkeypatch):
    monkeypatch.setattr("git_due_diligence.modules.security.shutil.which", lambda _: None)
    monkeypatch.setattr(cli, "_llm_unavailable_reason", lambda: None)

    def boom(prompt: str) -> str:
        raise RuntimeError("api down")

    monkeypatch.setattr(cli, "_anthropic_complete", boom)
    out = tmp_path / "r.md"
    result = runner.invoke(cli.app, ["analyze", str(fixture_repo), "--output", str(out), "--questions"])
    assert result.exit_code == 0
    md = out.read_text(encoding="utf-8")
    assert "Question for management" not in md
    assert "# Technical Due Diligence Report:" in md
