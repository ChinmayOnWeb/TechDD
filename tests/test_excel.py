from pathlib import Path

from openpyxl import load_workbook

from git_due_diligence.bridge.adjustments import Adjustment
from git_due_diligence.bridge.assumptions import Assumptions, CompCompany
from git_due_diligence.bridge.valuation import (
    blended_pre_dd_ev_mid,
    comps_valuation,
    dcf_scenarios,
    dcf_valuation,
    sensitivity_grid,
)
from git_due_diligence.excel import write_model


def _assumptions() -> Assumptions:
    return Assumptions(
        target_name="t",
        revenue_low=1_000_000, revenue_mid=2_000_000, revenue_high=3_000_000,
        revenue_source="press release",
        comps=[CompCompany("A", 4.0), CompCompany("B", 6.0), CompCompany("C", 8.0)],
        comps_source="filings",
        dcf_years=1,
        growth_bear=0.25, growth_base=0.50, growth_bull=0.75,
        operating_margin=0.20, discount_rate=0.25, terminal_growth=0.05,
        engineer_month_usd=20_000, retention_package_usd=300_000,
        integration_cost_usd=500_000,
        security_fix_cost_usd=50_000,
        license_discount_per_finding=0.02, license_discount_cap=0.10,
    )


def _adjustments() -> list[Adjustment]:
    return [
        Adjustment("Remediation capex", 20_000, 40_000, 60_000, "months x cost", ["hotspot: core/engine.py"]),
        Adjustment("Key-person retention", 300_000, 600_000, 900_000, "2 x package", ["dave"]),
        Adjustment("License-risk discount", 75_000, 150_000, 225_000, "2% of EV", ["mysqlclient"]),
        Adjustment("Integration cost", 250_000, 500_000, 750_000, "flat", []),
        Adjustment("Security remediation", 0, 0, 0, "Not assessed — security module ships in Phase 3", [], False),
    ]


def _write(tmp_path) -> Path:
    a = _assumptions()
    comps, dcf = comps_valuation(a), dcf_valuation(a)
    total_mid = sum(adj.mid for adj in _adjustments())
    path = tmp_path / "model.xlsx"
    write_model(path, a, _adjustments(), comps, dcf, dcf_scenarios(a),
                sensitivity_grid(a, total_mid))
    return path


def test_workbook_has_five_sheets_in_order(tmp_path):
    wb = load_workbook(_write(tmp_path))
    assert wb.sheetnames == ["Assumptions", "Comps", "DCF", "DD Adjustments", "Valuation Summary"]


def test_adjustments_sheet_rows_and_total_formula(tmp_path):
    ws = load_workbook(_write(tmp_path))["DD Adjustments"]
    assert ws["A2"].value == "Remediation capex"
    assert ws["C3"].value == 600_000
    assert ws["E6"].value == "NOT ASSESSED"
    assert ws["B7"].value == "=SUM(B2:B6)"
    assert ws["A7"].value == "Total"


def test_summary_formulas_and_disclaimer(tmp_path):
    ws = load_workbook(_write(tmp_path))["Valuation Summary"]
    assert ws["B4"].value == 6_000_000            # comps low
    assert ws["C5"].value == 3_000_000            # dcf mid
    assert ws["C6"].value == "=AVERAGE(C4:C5)"
    assert ws["C8"].value == "='DD Adjustments'!C7"
    assert ws["C9"].value == "=C6-C8"
    assert ws["B13"].value == "=$A13*B$12"        # pre-DD sensitivity cell
    assert ws["B19"].value == "=B13-$C$8"         # post-DD sensitivity cell
    assert "educational analysis" in ws["A23"].value.lower()


def test_assumptions_sheet_carries_sources(tmp_path):
    ws = load_workbook(_write(tmp_path))["Assumptions"]
    values = [ws.cell(row=r, column=3).value for r in range(1, 25)]
    assert "press release" in [v for v in values if v]


def test_assumptions_sheet_shows_every_pricing_input(tmp_path):
    """Every cost assumption actually used by the pricing bridge must be
    visible on the Assumptions tab, or the model can't be audited against
    its own inputs (security_fix_cost_usd was missing until this test)."""
    ws = load_workbook(_write(tmp_path))["Assumptions"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, 25)]
    assert "Security fix cost (USD)" in labels
    row = labels.index("Security fix cost (USD)") + 1
    assert ws.cell(row=row, column=2).value == 50_000
